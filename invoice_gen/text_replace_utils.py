import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from typing import List, Dict, Optional, Any, Union
import re
import datetime

# --- Helper to safely get nested data ---
def _get_nested_data(data: Dict, path: List[Union[str, int]]) -> Any:
    """Safely retrieves nested data using a list of keys/indices."""
    current_data = data
    try:
        for key_or_index in path:
            if isinstance(current_data, dict):
                current_data = current_data.get(key_or_index)
            elif isinstance(current_data, list) and isinstance(key_or_index, int):
                 if 0 <= key_or_index < len(current_data):
                     current_data = current_data[key_or_index]
                 else:
                     print(f"Warning: Index {key_or_index} out of bounds for list in path {path}")
                     return None # Index out of bounds
            else:
                print(f"Warning: Cannot traverse path {path} at step {key_or_index} with data type {type(current_data)}")
                return None # Cannot traverse further
            if current_data is None:
                # Allow reaching None without error, the caller should handle it.
                # print(f"Warning: Key/index {key_or_index} resulted in None in path {path}")
                return None # Key/index not found or led to None
        return current_data
    except (TypeError, IndexError, KeyError) as e:
        print(f"Error accessing data path {path}: {e}")
        return None

# --- NEW: Function to convert Excel numeric date to datetime ---
def excel_number_to_datetime(excel_num):
    """
    Convert an Excel date number to a Python datetime object.
    Excel dates are number of days since 1900-01-01.
    But Excel incorrectly thinks 1900 was a leap year, so we adjust for dates after Feb 28, 1900.
    """
    try:
        excel_num = float(excel_num)
        # Handle Excel's leap year bug (Excel thinks 1900 was a leap year)
        if excel_num > 59:  # if date is after February 28, 1900
            excel_num -= 1  # Adjust for the non-existent Feb 29, 1900
            
        # Calculate the datetime
        delta = datetime.timedelta(days=excel_num-1)  # Subtract 1 because Excel day 1 is 1900-01-01
        return datetime.datetime(1900, 1, 1) + delta
    except (ValueError, TypeError) as e:
        print(f"Error converting Excel date number: {e}")
        return None

# --- Function to detect date strings and format cells ---
def is_date_string(text: str) -> bool:
    """Check if a string resembles a date format."""
    # Common date patterns (can be expanded)
    date_patterns = [
        r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # MM/DD/YYYY, DD/MM/YYYY
        r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',    # YYYY/MM/DD
        r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{2,4}',  # DD Mon YYYY
        r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{2,4}'   # Mon DD, YYYY
    ]
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in date_patterns)

def is_excel_date_number(value: Any) -> bool:
    """Check if a value might be an Excel date number (integer between 1 and 2958465)."""
    try:
        # Convert to float and check range
        num_value = float(value)
        # Excel dates generally fall within this range (1900-01-01 to 9999-12-31)
        return 1 <= num_value <= 2958465
    except (ValueError, TypeError):
        return False

def format_cell_as_date(cell: Cell, value: Any) -> None:
    """
    Attempts to properly format a cell value as a date if possible.
    
    Args:
        cell: The Excel cell to format
        value: The value being inserted (string, datetime, or other)
    """
    # Check if it's a numeric value that could be an Excel date
    if is_excel_date_number(value):
        date_obj = excel_number_to_datetime(value)
        if date_obj:
            cell.value = date_obj
            cell.number_format = "dd/mm/yyyy"
            return
            
    # Always assign the value first
    cell.value = value
    
    # Check if it's a datetime object
    if isinstance(value, (datetime.datetime, datetime.date)):
        cell.number_format = "dd/mm/yyyy"  # Or any preferred format
        return
        
    # Check if it's a string that looks like a date
    if isinstance(value, str) and is_date_string(value):
        # Try to convert to Excel date format and display properly
        try:
            # Format the cell to display dates correctly regardless of internal value
            cell.number_format = "dd/mm/yyyy"  # Or any preferred format
        except Exception as e:
            print(f"Warning: Failed to format date string '{value}': {e}")
    
    # If neither condition is met, the cell remains with default formatting

# --- Modified function to replace a single value on one sheet ---
def find_and_replace_single_value(
    worksheet: Worksheet,
    text_to_find: str,
    replacement_value: Any,
    case_sensitive: bool = False,
    is_date: bool = False
):
    """
    Finds all occurrences of text_to_find in string cells and replaces them
    with the single replacement_value.
    """
    if text_to_find is None: return 0
    text_to_find_str = str(text_to_find)
    replacement_str = str(replacement_value)
    replacements_made = 0
    
    # Check if the replacement value might be an Excel date number
    excel_date_value = None
    if is_date and is_excel_date_number(replacement_value):
        excel_date_value = excel_number_to_datetime(replacement_value)
        if excel_date_value:
            print(f"Detected Excel date number {replacement_value}, converted to {excel_date_value.strftime('%Y-%m-%d')}")
    
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell): continue
            original_value = cell.value
            if isinstance(original_value, str):
                modified_value = None
                
                # :::::::::::::::::::: DEBUG: Check if placeholder might be in cell ::::::::::::::::::::
                should_check = False
                if case_sensitive and text_to_find_str in original_value:
                    should_check = True
                elif not case_sensitive and text_to_find_str.lower() in original_value.lower():
                    should_check = True

                if should_check:
                    print(f":::::::::::::::::::: Potential match for '{text_to_find_str}' found in {worksheet.title}!{cell.coordinate}. Original value: '{original_value}'")

                if case_sensitive:
                    if text_to_find_str in original_value:
                        modified_value = original_value.replace(text_to_find_str, replacement_str)
                else:
                    if text_to_find_str.lower() in original_value.lower():
                        new_value_parts = []
                        start_index = 0
                        find_len = len(text_to_find_str)
                        while start_index < len(original_value):
                            found_index = original_value.lower().find(text_to_find_str.lower(), start_index)
                            if found_index == -1: new_value_parts.append(original_value[start_index:]); break
                            else: new_value_parts.append(original_value[start_index:found_index]); new_value_parts.append(replacement_str); start_index = found_index + find_len
                        modified_value = "".join(new_value_parts)
                if modified_value is not None:
                    try: 
                        # Check if we need to format as date based on explicit flag or placeholder name
                        if is_date or text_to_find_str.lower() in ("jftime", "date", "invoice date"):
                            # For date cells, if the modified value is just the replacement value,
                            # we directly set the date value
                            if modified_value == replacement_str:
                                # First check if we have a pre-converted Excel date
                                if excel_date_value:
                                    cell.value = excel_date_value
                                    cell.number_format = "dd/mm/yyyy"
                                else:
                                    # Try to convert the date value to a proper Excel date
                                    try:
                                        # First attempt to parse the date
                                        if isinstance(replacement_value, (datetime.datetime, datetime.date)):
                                            date_obj = replacement_value
                                        else:
                                            # Try multiple date formats
                                            date_formats = [
                                                "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", 
                                                "%b %d, %Y", "%d %b %Y", "%B %d, %Y", "%d %B %Y"
                                            ]
                                            date_obj = None
                                            for fmt in date_formats:
                                                try:
                                                    date_obj = datetime.datetime.strptime(str(replacement_value).strip(), fmt)
                                                    break
                                                except ValueError:
                                                    continue
                                        
                                        if date_obj:
                                            # Successfully parsed as date, set as datetime object
                                            cell.value = date_obj
                                        else:
                                            # Couldn't parse as date, set as is and format
                                            cell.value = modified_value
                                        
                                        # Always apply date format to the cell
                                        cell.number_format = "dd/mm/yyyy"
                                    except Exception as date_err:
                                        print(f"Warning: Error formatting date '{replacement_value}': {date_err}")
                                        cell.value = modified_value
                            else:
                                # If it's a mixed text with the date, just set value and apply format
                                cell.value = modified_value
                                cell.number_format = "dd/mm/yyyy"
                        else:
                            cell.value = modified_value 
                        replacements_made += 1
                    except Exception as write_err: 
                        print(f"Warning: Error writing replaced value to {worksheet.title}!{cell.coordinate}: {write_err}")
    return replacements_made

# --- NEW Orchestrator for Data-Driven Replacements ---
def process_data_driven_replacements(
    workbook: openpyxl.Workbook,
    invoice_data: Dict[str, Any],
    replacement_rules: List[Dict[str, Any]]
):
    """
    Processes replacements where the value comes from the invoice_data dictionary.

    Args:
        workbook: The workbook to modify.
        invoice_data: The loaded invoice data dictionary.
        replacement_rules: List of rule dicts, e.g.,
            {
                "find": "PLACEHOLDER_TEXT",
                "data_path": ["processed_tables_data", "1", "po", 0], # Path to data
                "target_sheets": ["Invoice", "Contract"], # Sheets to apply to
                "case_sensitive": False # Optional
            }
    """
    print("\n--- Performing Data-Driven Replacements ---")
    if not replacement_rules:
        print("DEBUG: No data-driven replacement rules provided.")
        return

    total_replacements_count = 0
    for rule in replacement_rules:
        placeholder = rule.get("find")
        data_path = rule.get("data_path")
        target_sheets = rule.get("target_sheets")
        case_sensitive = rule.get("case_sensitive", False)
        is_date = rule.get("is_date", False)  # Flag to indicate date handling

        if not placeholder or not data_path or not target_sheets:
            print(f"Warning: Skipping invalid data-driven rule: {rule}")
            continue

        # Retrieve the replacement value from invoice_data using the path
        replacement_value = _get_nested_data(invoice_data, data_path)

        # :::::::::::::::::::: DEBUG: Print fetched value ::::::::::::::::::::
        print(f":::::::::::::::::::: Rule '{placeholder}' - Fetched Value: '{replacement_value}' (Type: {type(replacement_value)}) from path {data_path}")

        # --- Enhanced Logging for Data Fetching ---
        # Removed original enhanced logging block as the new debug print above covers it.
        # --- End Enhanced Logging ---

        if replacement_value is None:
            print(f"Warning: Could not find data at path {data_path} for placeholder '{placeholder}'. Skipping this rule.")
            continue

        print(f"Attempting replacement for '{placeholder}' with value '{replacement_value}' on sheets: {target_sheets}")
        if is_date:
            print(f"  This is a DATE field and will be formatted accordingly")

        rule_replacements_count = 0
        for sheet_name in target_sheets:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                # Call the other utility function to do the actual replacement on this sheet
                count = find_and_replace_single_value(
                    worksheet=worksheet,
                    text_to_find=placeholder,
                    replacement_value=replacement_value,
                    case_sensitive=case_sensitive,
                    is_date=is_date  # Pass the date flag to the replacement function
                )
                rule_replacements_count += count
            else:
                print(f"Warning: Target sheet '{sheet_name}' not found for placeholder '{placeholder}'.")

        if rule_replacements_count > 0:
             print(f"Made {rule_replacements_count} replacement(s) across sheets for placeholder '{placeholder}'.")
        total_replacements_count += rule_replacements_count

    print(f"--- Finished Data-Driven Replacements. Total replacements made: {total_replacements_count} ---")


# Keep find_and_replace_in_workbook if you still need it for config-driven global replacements
def find_and_replace_in_workbook(
    workbook: openpyxl.Workbook,
    replacement_rules: List[Dict[str, Any]],
    target_sheets: Optional[List[str]] = None
):
    """
    Performs find-and-replace operations on specified sheets within a workbook.

    Args:
        workbook: The openpyxl Workbook object to modify.
        replacement_rules: A list of dictionaries, where each dictionary defines a rule:
            {
                "find": "text_to_find",
                "replace": "replacement_text",
                "case_sensitive": False (Optional, defaults to False)
                # Add future options like 'whole_word' if needed
            }
        target_sheets: Optional list of sheet names to process. If None, processes all visible sheets.
    """
    if not replacement_rules:
        print("DEBUG: No replacement rules provided. Skipping text replacement.")
        return

    print(f"--- Starting Global Text Replacement ---")

    sheets_to_process = []
    if target_sheets:
        sheets_to_process = [sheet_name for sheet_name in target_sheets if sheet_name in workbook.sheetnames]
        print(f"Targeting specific sheets for replacement: {sheets_to_process}")
    else:
        sheets_to_process = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
        print(f"Targeting all visible sheets for replacement: {sheets_to_process}")

    if not sheets_to_process:
        print("Warning: No valid sheets found to process for replacements.")
        return

    total_replacements_made = 0

    for sheet_name in sheets_to_process:
        worksheet = workbook[sheet_name]
        print(f"Processing sheet: '{sheet_name}' for replacements...")
        sheet_replacements_made = 0

        # Iterate through all cells in the sheet
        # Using iter_rows() is generally efficient
        for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                # Skip merged cells (only process the top-left origin cell implicitly)
                if isinstance(cell, MergedCell):
                    continue

                original_value = cell.value

                # Only process cells containing strings
                if isinstance(original_value, str):
                    current_value = original_value
                    cell_modified = False
                    for rule in replacement_rules:
                        text_to_find = rule.get("find")
                        replacement_text = rule.get("replace", "") # Default to empty string if missing
                        case_sensitive = rule.get("case_sensitive", False)

                        if text_to_find is None: # Skip rule if 'find' is missing
                            continue

                        # Ensure find/replace values are strings
                        text_to_find = str(text_to_find)
                        replacement_text = str(replacement_text)

                        # Perform replacement
                        if case_sensitive:
                            if text_to_find in current_value:
                                current_value = current_value.replace(text_to_find, replacement_text)
                                cell_modified = True
                        else:
                            # Simple case-insensitive replace (more complex regex could be used for edge cases)
                            if text_to_find.lower() in current_value.lower():
                                # Rebuild string to handle case-insensitivity (simple approach)
                                # This might not be perfect for overlapping matches but covers many cases
                                new_value_parts = []
                                start_index = 0
                                find_len = len(text_to_find)
                                while start_index < len(current_value):
                                    found_index = current_value.lower().find(text_to_find.lower(), start_index)
                                    if found_index == -1:
                                        new_value_parts.append(current_value[start_index:])
                                        break
                                    else:
                                        new_value_parts.append(current_value[start_index:found_index])
                                        new_value_parts.append(replacement_text)
                                        start_index = found_index + find_len
                                current_value = "".join(new_value_parts)
                                cell_modified = True

                    # If any rule modified the cell, write the final value back
                    if cell_modified:
                        try:
                            cell.value = current_value
                            sheet_replacements_made += 1
                            # print(f"DEBUG: Replaced in {sheet_name}!{cell.coordinate}: '{original_value}' -> '{current_value}'")
                        except Exception as write_err:
                             print(f"Warning: Error writing replaced value to {sheet_name}!{cell.coordinate}: {write_err}")

        if sheet_replacements_made > 0:
            print(f"Made {sheet_replacements_made} replacement(s) in sheet '{sheet_name}'.")
        total_replacements_made += sheet_replacements_made

    print(f"--- Finished Global Text Replacement. Total replacements made: {total_replacements_made} ---")
