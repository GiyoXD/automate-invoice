import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
# from openpyxl.worksheet.dimensions import RowDimension # Not strictly needed for access
from typing import Dict, List, Optional, Tuple, Any

# --- store_original_merges FILTERED to ignore merges ABOVE row 16 ---
def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float], int]]]:
    """
    Stores the horizontal span (colspan), the value of the top-left cell,
    the height of the starting row, and the starting column index for
    merged ranges in specified sheets that start at row 16 or below.
    This version has robust, separated error handling for height and value.
    """
    original_merges = {}
    print("\nStoring original merge data (span, value, height, and start column)...")
    print("  (Ignoring merges that start above row 16)")

    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name]
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_count = 0

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                if max_row != min_row or min_row < 16:
                    skipped_count += 1
                    continue

                col_span = max_col - min_col + 1
                row_height = None  # Default to None
                top_left_value = None # Default to None

                # --- START: MODIFIED ERROR HANDLING ---
                # Step 1: Try to get the row height.
                try:
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                except KeyError:
                    # This is not an error, it just means the row has default height.
                    # row_height remains None, which is correct.
                    pass
                except Exception as e:
                    print(f"    Warning: Could not get row height for row {min_row}. Storing as None. Error: {e}")

                # Step 2: Try to get the cell value.
                try:
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value
                except Exception as e:
                    print(f"    Warning: Could not get cell value at ({min_row},{min_col}). Storing as None. Error: {e}")
                
                # Step 3: Append all the data we gathered.
                merges_data.append((col_span, top_left_value, row_height, min_col))
                # --- END: MODIFIED ERROR HANDLING ---

            original_merges[sheet_name] = merges_data
            print(f"  Stored {len(merges_data)} merge data entries for sheet '{sheet_name}'.")
            if skipped_count > 0:
                print(f"    (Skipped {skipped_count} irrelevant merges)")
        else:
             print(f"  Warning: Sheet '{sheet_name}' not found during merge storage.")
             original_merges[sheet_name] = []
             
    return original_merges

# --- find_and_restore_merges_heuristic remains unchanged (still searches bottom-up, applies stored value/height) ---
# In merge_utils.py

def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float], int]]],
                                      processed_sheet_names: List[str]):
    """
    Restores merges by searching for a stored value within its original starting column,
    using the sheet's actual max row and column as search boundaries.
    This version uses a manual overlap check for compatibility with older openpyxl versions.
    """
    print("Starting accurate merge restoration process...")
    restored_count = 0
    failed_count = 0

    for sheet_name in processed_sheet_names:
        if sheet_name not in workbook.sheetnames or sheet_name not in stored_merges:
            continue

        worksheet: Worksheet = workbook[sheet_name]
        original_merges_data = stored_merges[sheet_name]
        successfully_restored_values_on_sheet = set()

        search_min_row = 16
        search_max_row = worksheet.max_row
        
        print(f"  Processing sheet '{sheet_name}', searching up to row {search_max_row}.")

        for col_span, stored_value, stored_height, original_start_col in original_merges_data:
            # ... (code for skipping rules) ...

            found = False
            for r in range(search_max_row, search_min_row - 1, -1):
                current_cell = worksheet.cell(row=r, column=original_start_col)

                if current_cell.value == stored_value:
                    start_row, start_col = r, original_start_col
                    end_row = start_row
                    end_col = start_col + col_span - 1
                    
                    try:
                        # --- START: MODIFIED OVERLAP CHECK ---
                        # This block replaces the .overlaps() method for compatibility.
                        for existing_merge in list(worksheet.merged_cells.ranges):
                            rows_overlap = (existing_merge.min_row <= end_row) and (existing_merge.max_row >= start_row)
                            cols_overlap = (existing_merge.min_col <= end_col) and (existing_merge.max_col >= start_col)
                            
                            if rows_overlap and cols_overlap:
                                worksheet.unmerge_cells(str(existing_merge))
                        # --- END: MODIFIED OVERLAP CHECK ---

                        worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                        # ... (rest of the try block) ...
                        
                        restored_count += 1
                        found = True
                        break

                    except Exception as e:
                        # This is the line that printed your previous errors
                        print(f"    Warning: Failed to apply merge for value '{str(stored_value)[:50]}' at ({start_row},{start_col}). Error: {e}")
                        failed_count += 1
                        found = True 
                        break

            if not found and stored_value not in successfully_restored_values_on_sheet:
                failed_count += 1

    print(f"Merge restoration finished. Restored: {restored_count}, Failed: {failed_count}.")